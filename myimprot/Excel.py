# -*- coding: utf-8 -*-
# @Author:S
# @Time    : 2022/8/11 03:21


import openpyxl


class execl_op:
    # 绑定
    def __init__(self, fnmae):
        self.fname = fname
        self.wb = openpyxl.load_workbook(self.fname)
        sheets = self.wb.sheetnames
        self.sheet = sheets[0]
        self.ws = self.wb[self.sheet]

    # 获取表格的总行数和总列数
    def get_row_clo_num(self):
        rows = self.ws.max_row
        columns = self.ws.max_column
        return rows, columns

    # 获取表格某个单元格的值
    def get_cell_value(self, row, column):
        cell_value = self.ws.cell(row=row, column=column).value
        return cell_value

    # 获取某列所有值
    def get_col_value(self, column):
        rows = self.ws.max_row
        column_data = []
        for i in range(1, rows + 1):
            cell_value = self.ws.cell(row=i, column=column).value
            column_data.append(cell_value)
        return column_data

    # 获取某行所有值
    def get_row_value(self, row):
        columns = self.ws.max_column
        row_data = []
        for i in range(1, columns + 1):
            cell_value = self.ws.cell(row=row, column=i).value
            row_data.append(cell_value)
        return row_data

    # 根据单元格的值获取列号
    def get_cell_column(self):
        column = self.ws.max_column
        for i in range(1, column + 1):
            cell_value = self.ws.cell(row=1, column=i).value
            if cell_value == 'SN' or cell_value == 'sn':
                return i

    # 设置某个单元格的值
    def set_cell_value(self, row, colum, cellvalue):
        def set_cell_value(self, row, colunm, cellvalue):
            try:
                self.ws.cell(row=row, column=colunm).value = cellvalue
                self.wb.save(self.file)
            except:
                self.ws.cell(row=row, column=colunm).value = "writefail"
                self.wb.save(self.file)


if __name__ == '__main__':
    num = input("请输入多少台设备：")
    fname = "E:\ks\dv%s.xlsx" % num
    first_row = execl_op(fname).get_row_value(1)
    print(first_row)
    second_col = execl_op(fname).get_col_value(5)
    print(second_col)
    column_num = execl_op(fname).get_cell_column()
    print('SN的列数是：%s' % column_num)
